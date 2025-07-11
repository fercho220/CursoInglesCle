from django.db import IntegrityError
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import redirect, render
from .forms import *
from .models import *
from django.views.generic import View,TemplateView, ListView, UpdateView, CreateView, DeleteView
from django.urls import reverse, reverse_lazy
from django.contrib.auth.models import User
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth import authenticate , login, logout
from usuario.mixins import *
from django.forms.models import modelformset_factory
from django.db.models import Q
# librerias para generar pdf
import os
from django.conf import settings
from django.template.loader import get_template
from django.template import Context
# from xhtml2pdf import pisa
from weasyprint import HTML, CSS
# from ingles.utils import render_to_pdf
from datetime import datetime

# librerias para generar excel
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
import cairosvg
import datetime

#librerias para respuestas
from django.http import HttpResponseBadRequest
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.shortcuts import render, redirect
from django.contrib import messages
from django.urls import reverse
from django.views.generic import UpdateView
from django.http import JsonResponse


#corrección pago
from django.http import HttpResponseBadRequest

from django.shortcuts import render
from django.contrib.messages import get_messages




## MENSAJES
def fetch_messages(request):
    storage = get_messages(request)
    return render(request, 'ingles/message.html', {'messages': storage})

def fetch_comments_status(request):
    pagos = Pago.objects.all().values('foliopago', 'descripcion', 'idestado__estado')
    comments_status = [
        {
            'foliopago': pago['foliopago'],
            'has_comments': bool(pago['descripcion'].strip()),
            'estado': pago['idestado__estado']
        }
        for pago in pagos
    ]
    return JsonResponse({'comments_status': comments_status})

##---------------------FUNCIONAL_ESTADISTICAS_INDIVIDUALES----------------##
##    path('estadisticas/<int:pk>', Estadisticas.as_view(), name = 'estadisticas' ),
class Estadisticas( View ):

    model = Grupo
    template_name = 'estadisticas/Estadistica_Individual.html'

    def get_queryset(self,pk):
        return self.model.objects.get(idgrupo=pk)

    def get(self, request,pk, *args, **kwargs):
        opp = Grupo.objects.get(idgrupo=pk)

        # grupos = Grupo.objects.filter(idperiodo=idperiodo).filter(idmateria=idmateria).filter(estado=True)
        # print(grupos)

        #MODELS Carrera # idcarrera # nombrecarrera
        gruposdetalles = Det_Grupo.objects.filter(
            idgrupo=pk).order_by('idestudiante__apellidop')
        var = Det_Grupo.objects.filter(idgrupo=pk).count()  
        # 26 inscritos
        group=Det_Grupo.objects.filter(idgrupo=pk).filter(foliopago__idestado=2).count() #varA 25 aceptados
        abre=Det_Grupo.objects.filter(idgrupo=opp).filter(foliopago__idestado=2)

        carreritas = Carrera.objects.all()
        nombrecarreras = ""     #str
        idcarreras = ""         #str 1,2,3,6,7,4,8,9,10,
        cont_id=0 #10 equivale a 1 2 3 6 7 4 8 9 10 

        data_idcarreras="" #7,6,1,3,2,1,4,2,2,0,0,0,2,2,0,7,3,4,2,1,1,0,0,0,0,0,0, #abajo esta +apro+repro
        group_male=0 #16
        group_female=0 #9
        # list=[]
        for elemento in carreritas:

            cont_id=elemento.idcarrera
            race = abre.filter(idestudiante__idcarrera= cont_id).count()
            if race > 0:
                # nombrecarreras=nombrecarreras +"'"+ elemento.nombrecarrera +"',"
                nombrecarreras=nombrecarreras +"'"+ elemento.nombrecarrera +"',"+("'Hombres-"+elemento.nombrecarrera)+"',"+("'Mujeres-"+elemento.nombrecarrera)+"',"
                idcarreras = idcarreras + str(elemento.idcarrera) + ","
                male = abre.filter(idestudiante__idcarrera = cont_id).filter(idestudiante__genero="M").count()             
                female = abre.filter(idestudiante__idcarrera = cont_id).filter(idestudiante__genero="F").count()
                data_idcarreras = data_idcarreras + str(race) + ","+ str(male) + ","+ str(female) + ","

                group_male = group_male+male
                group_female = group_female+female

        abre2=Det_Grupo.objects.filter(idgrupo = pk)
        aprobado=0 #17 int 
        for c2 in range(70,101): ## cuenta del 7-10 range(7)
            var_aprobado = abre2.filter(calif=c2).count()
            aprobado = aprobado + var_aprobado

        reprobado=0
        for c3 in range(0,69): ## cuenta del 0al6 range(7)
            var_reprobado = abre2.filter(calif=c3).count()
            reprobado = reprobado + var_reprobado
            # print(var_reprobado)
        
        NA = Det_Grupo.objects.filter(idgrupo = pk).filter(calif="NA").count()
        zero = Det_Grupo.objects.filter(idgrupo = pk).filter(calif="00").count()
        reprobado = reprobado +NA + zero

        nombrecarreras=nombrecarreras +"'Aprobados',"+"'Reprobados'," + \
            "'Grupo',"+"'Grupo-Hombres'," +"'Grupo-Mujeres'," 
#Se agrego los valores faltantes
        valo = data_idcarreras + \
            str(aprobado) + "," + str(reprobado) + "," + \
            str(group) + "," + str(group_male) + "," + str(group_female) + ","

        context = { 
            'valo':valo, 
            'nombrecarreras':nombrecarreras, 
            'carreritas':carreritas,
            'gruposdetalles':gruposdetalles,
            'var':var,
            'group': group,
            'idcarreras':idcarreras,
        }
      
        return render(request, self.template_name, context,) 

class EstadisticasGlobales( View ):

    model = Grupo
    template_name = 'estadisticas/Estadistica_Global.html'
    form_class = PeriodoForm 

    def get_queryset(self,pk):
        return self.model.objects.get(idperiodo=pk)

    def get(self, request,pk, *args, **kwargs):
        # grupos = Grupo.objects.filter(idperiodo=idperiodo).filter(idmateria=idmateria).filter(estado=True)
        # print(grupos)

        #MODELS Carrera # idcarrera # nombrecarrera
        if pk == 'all':
            pk = Periodo.objects.latest('idperiodo')
            
        gruposdetalles = Det_Grupo.objects.filter(
            idperiodo=pk).order_by('idestudiante__apellidop')
        var = Det_Grupo.objects.filter(idperiodo=pk).count()  
        # 26 inscritos
        group=Det_Grupo.objects.filter(idperiodo=pk).filter(foliopago__idestado=2).count() #varA 25 aceptados
        abre=Det_Grupo.objects.filter(idperiodo=pk).filter(foliopago__idestado=2)

        carreritas = Carrera.objects.all()
        nombrecarreras = ""     #str
        idcarreras = ""         #str 1,2,3,6,7,4,8,9,10,
        cont_id=0 #10 equivale a 1 2 3 6 7 4 8 9 10 

        data_idcarreras="" #7,6,1,3,2,1,4,2,2,0,0,0,2,2,0,7,3,4,2,1,1,0,0,0,0,0,0, #abajo esta +apro+repro
        group_male=0 #16
        group_female=0 #9
        group_none=0 #9
        # list=[]
        carreras = []
        totales = []
        total = 0
        for elemento in carreritas:

            cont_id=elemento.idcarrera
            race = abre.filter(idestudiante__idcarrera= cont_id).count()
            if race > 0:
                # nombrecarreras=nombrecarreras +"'"+ elemento.nombrecarrera +"',"
                nombrecarreras=nombrecarreras +"'"+ elemento.nombrecarrera +"',"+("'Hombres-"+elemento.nombrecarrera)+"',"+("'Mujeres-"+elemento.nombrecarrera)+"',"
                idcarreras = idcarreras + str(elemento.idcarrera) + ","
                male = abre.filter(idestudiante__idcarrera = cont_id).filter(idestudiante__genero="M").count()             
                female = abre.filter(idestudiante__idcarrera = cont_id).filter(idestudiante__genero="F").count()
                nane = abre.filter(idestudiante__idcarrera=cont_id).exclude(Q(idestudiante__genero="M") | Q(idestudiante__genero="F")).count()
                data_idcarreras = data_idcarreras + str(race) + ","+ str(male) + ","+ str(female) + ","

                group_male = group_male+male
                group_female = group_female+female
                group_none  = group_none+nane

                carreras.append(elemento.nombrecarrera)
                totales.append(race)
                total = total + race        

        abre2=Det_Grupo.objects.filter(idperiodo = pk)
        niveles=[]
        totalNivel = []
        n = 1
        while n != 6:
            nivel = abre.filter(idgrupo__idmateria__nombremateria='NIVEL ' + str(n)).count()
            niveles.append('Nivel ' + str(n))
            totalNivel.append(nivel)
            n = n + 1

        aprobado=0 #17 int 
        for c2 in range(70,101): ## cuenta del 7-10 range(7)
            var_aprobado = abre2.filter(calif=c2).count()
            aprobado = aprobado + var_aprobado

        reprobado=0
        for c3 in range(0,69): ## cuenta del 0al6 range(7)
            var_reprobado = abre2.filter(calif=c3).count()
            reprobado = reprobado + var_reprobado
            # print(var_reprobado)
        
        NA = Det_Grupo.objects.filter(idperiodo = pk).filter(calif="NA").count()
        zero = Det_Grupo.objects.filter(idperiodo = pk).filter(calif="00").count()
        reprobado = reprobado +NA + zero

        nombrecarreras=nombrecarreras +"'Aprobados',"+"'Reprobados'," + \
            "'Alumnos',"+"'Alumnos-Hombres'," +"'Alumnos-Mujeres'," 
#Se agrego los valores faltantes
        valo = data_idcarreras + \
            str(aprobado) + "," + str(reprobado) + "," + \
            str(group) + "," + str(group_male) + "," + str(group_female) + ","
        
        periodo = Periodo.objects.order_by('-idperiodo')
        periodos = []
        for aao in periodo:
            periodos.append(aao)


        
        context = { 
            'valo':valo, 
            'nombrecarreras':nombrecarreras, 
            'carreritas':carreritas,
            'gruposdetalles':gruposdetalles,
            'var':var,
            'group': group,
            'idcarreras':idcarreras,
            'periodos':periodos,
            'group_male':group_male,
            'group_female':group_female,
            'group_none':group_none,
            'aprobado': aprobado,
            'reprobado': reprobado,
            'totales':totales,
            'carrera':carreras,
            'niveles':niveles,
            'totalniveles':totalNivel,
            'total':total
        }
      
        return render(request, self.template_name, context,) 

# CICLOS
# 'ISC','Male-ISC','female-ISC','IEM','Male-IEM','female-IEM','IBQ','Male-IBQ','female-IBQ','LA','Male-LA','female-LA','CP','Male-CP','female-CP','ARQ','Male-ARQ','female-ARQ','IGE','Male-IGE','female-IGE','MSC','Male-MSC','female-MSC','EXT','Male-EXT','female-EXT','Aprobados','Reprobados', 
    # ADD, grupo, grupo-hombres, grupo-mujeres
#valo 7,6,1,3,2,1,4,2,2,0,0,0,2,2,0,7,3,4,2,1,1,0,0,0,0,0,0,17,8,
    # ADD 25,16,9

# idcarreras estan en str
#   1,2,3,6,7,4,8,9,10, estos son los id estan en str
 # SE MODOFICO highcharts.js
# agrego unicamente el super user lo visualice listar_grupos.html

##---------------------ENDFUNCIONAL_ESTADISTICAS_INDIVIDUALES----------------##

##---------------------PDF_ESTADISTICAS_PROCESO!!----------------##
#    path('grafica_pdf/<int:pk>', Grafic_pdf.as_view(), name = 'grafica_pdf'),
class Grafic_pdf(LoginYSuperUsuarioMixin,View):
    def get(self, request,*args,**kwargs):
        try:
            template = get_template('estadisticas/imprimir.html')

            group=Det_Grupo.objects.filter(idgrupo = self.kwargs['pk']).filter(foliopago__idestado=2).count()

            abre=Det_Grupo.objects.filter(idgrupo=self.kwargs['pk']).filter(foliopago__idestado=2)
            carreritas = Carrera.objects.all()
            nombrecarreras = ""     #str
            idcarreras = ""         #str 1,2,3,6,7,4,8,9,10,
            cont_id=0 #10 equivale a 1 2 3 6 7 4 8 9 10 
            data_idcarreras="" #7,6,1,3,2,1,4,2,2,0,0,0,2,2,0,7,3,4,2,1,1,0,0,0,0,0,0, #abajo esta +apro+repro
            group_male=0 #16
            group_female=0 #9
            
            for elemento in carreritas:
                cont_id=elemento.idcarrera
                race = abre.filter(idestudiante__idcarrera= cont_id).count()
                if race > 0:
                    # nombrecarreras=nombrecarreras +"'"+ elemento.nombrecarrera +"',"
                    nombrecarreras=nombrecarreras +"'"+ elemento.nombrecarrera +"',"+("'Hombres-"+elemento.nombrecarrera)+"',"+("'Mujeres-"+elemento.nombrecarrera)+"',"
                    idcarreras = idcarreras + str(elemento.idcarrera) + ","
                    male = abre.filter(idestudiante__idcarrera = cont_id).filter(idestudiante__genero="M").count()             
                    female = abre.filter(idestudiante__idcarrera = cont_id).filter(idestudiante__genero="F").count()
                    data_idcarreras = data_idcarreras + str(race) + ","+ str(male) + ","+ str(female) + ","

                    group_male = group_male+male
                    group_female = group_female+female

            abre2=Det_Grupo.objects.filter(idgrupo = pk)
            aprobado=0 #17 int 
            for c2 in range(70,101): ## cuenta del 7-10 range(7)
                var_aprobado = abre2.filter(calif=c2).count()
                aprobado = aprobado + var_aprobado

            reprobado=0
            for c3 in range(0,69): ## cuenta del 0al6 range(7)
                var_reprobado = abre2.filter(calif=c3).count()
                reprobado = reprobado + var_reprobado
                # print(var_reprobado)
            
            NA = Det_Grupo.objects.filter(idgrupo = pk).filter(calif="NA").count()
            zero = Det_Grupo.objects.filter(idgrupo = pk).filter(calif="00").count()
            reprobado = reprobado +NA + zero

            # #Se agrego los nombres faltantes
            nombrecarreras=nombrecarreras +"'Aprobados',"+"'Reprobados'," + \
            "'Grupo',"+"'Grupo-Hombres'," +"'Grupo-Mujeres'," 
            #Se agrego los valores faltantes
            valo = data_idcarreras + str(aprobado)+"," + str(reprobado) + "," + str(group) + "," + str(group_male) + "," + str(group_female) + ","
            print (valo)
            print(type(valo))

            context = {
                'valo':valo,
                'group':group,

                'aprobado':aprobado, 
                'ico' : '{}{}'.format(settings.MEDIA_URL, 'TecNMHorizontal1.png'),
                # gruposdetalles = Det_Grupo.objects.filter(idgrupo=pk).order_by('idestudiante__apellidop')
                'sale': Det_Grupo.objects.filter(idgrupo=self.kwargs['pk']).order_by('idestudiante__apellidop'),
                # var = Det_Grupo.objects.filter(idgrupo=pk).count()  # 26
                'var': Det_Grupo.objects.filter(idgrupo = self.kwargs['pk']).count(), # 26
                
                
                'nombrecarreras':nombrecarreras, 
                # 'valo':valo,
                #gruposdetalles es sale

            }
            html = template.render(context)
            css_url = os.path.join(settings.BASE_DIR, 'static/css/bootstrap.min.css')
            pdf = HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(stylesheets=[CSS(css_url)])
            return HttpResponse(pdf, content_type='application/pdf')
        except:
            pass
        return HttpResponseRedirect(reverse_lazy('ingles:listar_grupos'))


##---------------------END_PDF_ESTADISTICAS_PROCESO!!----------------##



# Create your views here.   
# <td> <button href="{% url 'http://127.0.0.1:8000/media/' pago.pagocurso %}"  class="btn btn-outline-primary">ver</button></td>
def home(request):  
    if request.user.is_authenticated:
        user1 = request.user
        user_groups = user1.groups.values_list("name", flat=True)
        print(user_groups)
        if 'Facilitadores' in user_groups:
            print('si hay')
            facilitador = User.objects.get(username = request.user) or Docente.objects.get(usuario = request.user)
            return render(request, 'indexF.html', {'facilitador': facilitador} )
    return render(request, 'index.html')

def home2(request):  
    if request.user.is_authenticated:
        user1 = request.user
        user_groups = user1.groups.values_list("name", flat=True)
        print(user_groups)
        if 'Facilitadores' in user_groups:
            print('si hay')
            facilitador = User.objects.get(username = request.user) or Docente.objects.get(usuario = request.user)
            return render(request, 'indexF.html', {'facilitador': facilitador} )
    return render(request, 'index.html')
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""

class adminP(TemplateView):
    template_name = 'home/admin.html'

class Inicio(LoginRequiredMixin,TemplateView):
    template_name = 'index.html'
    
    def home(request):  
        if request.user.is_authenticated:
            estudiante = Estudiante.objects.get(usuario = request.user)
            return render(request, 'index.html', {'estudiante': estudiante} )
        return render(request, 'index.html')

class Inicio2(LoginRequiredMixin,TemplateView):
    template_name = 'home/index.html'
    
    def home(request):  
        if request.user.is_authenticated:
            estudiante = Estudiante.objects.get(usuario = request.user)
            return render(request, 'home/index.html', {'estudiante': estudiante} )
        return render(request, 'home/index.html')
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""

def Registro(request):
    data = {
        'form': CustomUserCreationForm(),
        'form2': CarreraForm()
    }
    if request.method=='POST':
        formulario = CustomUserCreationForm(data = request.POST)
        formulario2 = CarreraForm(data = request.POST)
        if formulario.is_valid():
            formulario.save()
            formu = formulario2.save(commit = False)
            user = authenticate(request, username=formulario.cleaned_data["username"], password=formulario.cleaned_data["password1"])
            try: 
                alumno = Estudiante.objects.create(nocontrol=formulario.cleaned_data["username"],nombre=formulario.cleaned_data["first_name"],apellidop=formulario.cleaned_data["last_name"] ,email=formulario.cleaned_data["email"],usuario=user,idcarrera=formu.idcarrera,apellidom=formu.apellidom,genero = formu.genero)
            except 	IntegrityError:
                user.delete()
                return render(request, 'ingles/registro.html',{'error_message':'Ya Existe El Correo Electronico'})
            login(request, user)
            return HttpResponseRedirect(reverse('ingles:index'))
        data['form'] = formulario
    return render(request, 'ingles/registro.html', data)
    
class PerfilListadoEstudiante(LoginRequiredMixin,View):
    model = Estudiante
    modelg = Det_Grupo
    form_class = EstudianteForm 
    template_name = 'ingles/listar_perfil.html'
    def get_queryset(self):
        return self.model.objects.filter(usuario = self.request.user)

    def get_context_data(self, **kwargs):
        contexto = {}
        list = []
        for alumno in self.get_queryset():
            gTotal = self.modelg.objects.filter(idestudiante = alumno.idestudiante)
            for grupo in gTotal:
                try:
                    calif = int(grupo.calif)
                    if calif < 70:
                        grupo.aprobado = 2
                    elif calif > 100:
                        grupo.aprobado = 4
                    else:
                        grupo.aprobado = 1
                except (ValueError, TypeError):
                    if grupo.calif == 'NA':
                        grupo.aprobado = 2
                    else:
                        grupo.aprobado = 3 
                list.append({'alumno':alumno,'grupos':grupo})
        contexto['estudiantes']=list
        return contexto
    
    def get(self, request, *args, **kwargs):    
        return render(request, self.template_name, self.get_context_data())




class PerfilListadoFacilitador(LoginRequiredMixin,LoginYFacilitador,View):
    model = Docente
    form_class = DocenteForm
    template_name = 'ingles/listar_perfilF.html'
    def get_queryset(self):
        return self.model.objects.filter(usuario = self.request.user)
    
    def get_context_data(self, **kwargs):
        contexto = {}
        contexto['Facilitador']=self.get_queryset()
        return contexto
    
    def get(self, request, *args, **kwargs):    
        return render(request, self.template_name, self.get_context_data())
    
    # def home(request):  
    #     if request.user.is_authenticated:
    #         estudiante = Estudiante.objects.get(usuario = request.user)
    #         return render(request, 'ingles/crear_pago.html', {'estudiante': estudiante} )
    #     return render(request, 'ingles/crear_pago.html')

class PerfilActualizarEstudiante(LoginRequiredMixin, UpdateView):
    model = Estudiante
    form_class = EstudianteForm
    template_name = 'ingles/perfil.html'
    success_url = reverse_lazy('curso:listar_perfil')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['estudiantes']=Estudiante.objects.filter(estado = True)
        return context

class PerfilActualizarFacilitador(LoginRequiredMixin,LoginYFacilitador, UpdateView):
    model = Docente
    form_class = FacilitadorForm
    template_name = 'ingles/perfilf.html'
    success_url = reverse_lazy('curso:listar_perfilF')
    # success_message = " was created successfully"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['Facilitador']=Docente.objects.filter(estado = True)
        return context
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
""" CRUD ESTUDIANTE """
class ListadoEstudiante(LoginYSuperUsuarioMixin, View):
    model = Estudiante
    form_class = EstudianteForm 
    template_name = 'ingles/listar_estudiante.html'

    def get_queryset(self):
        return self.model.objects.filter(estado = True)
    
    def get_context_data(self, **kwargs):
        contexto = {}
        contexto['estudiantes']=self.get_queryset()
        return contexto
    
    def get(self, request, *args, **kwargs):    
        return render(request, self.template_name, self.get_context_data())
    
class ActualizarEstudiante(LoginYSuperUsuarioMixin, UpdateView):
    model = Estudiante
    form_class = EstudianteForm
    template_name = 'ingles/estudiante.html'
    success_url = reverse_lazy('curso:listar_estudiante')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['estudiantes']=Estudiante.objects.filter(estado = True)
        return context

class CrearEstudiante(LoginYSuperUsuarioMixin, CreateView):
    model = Estudiante
    template_name = 'ingles/crear_estudiante.html'
    form_class = EstudianteForm
    success_url = reverse_lazy('curso:listar_estudiante')


class EliminarEstudiante(LoginYSuperUsuarioMixin, DeleteView):
    model = Estudiante
    #success_url = reverse_lazy('curso:listar_estudiante')

    def post(self,request,pk, *args, **kwargs):
        object = Estudiante.objects.get(idestudiante = pk)
        object.estado = False
        object.save()
        return redirect('curso:listar_estudiante')
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
    
""" CRUD GRUPO """
class ListadoGrupo(LoginRequiredMixin, View):
    model = Grupo
    form_class = GruposForm
    template_name = 'ingles/listar_grupos.html'

    def get_queryset(self):
        # return self.model.objects.all().order_by('idmateria')
        return self.model.objects.filter(estadoG=True).order_by('idperiodo','iddocente')

    
    def get_context_data(self, **kwargs):
        # var = Det_Grupo.objects.filter(idgrupo =self.model.objects.all().idgrupo ).count()
        # print(var)
        contexto = {}
        # contexto['grupos']=self.get_queryset()
        list=[]
        for grupo in self.get_queryset():
            var = Det_Grupo.objects.filter(idgrupo =grupo.idgrupo).count()
            list.append({'grupos':grupo,'counter':var})
        contexto['var']=list

        # print(contexto)
        return contexto
    def get(self, request, *args, **kwargs):    
        return render(request, self.template_name, self.get_context_data())

class ListadoGrupoF(LoginRequiredMixin, View):
    model = Grupo
    form_class = GruposForm
    template_name = 'ingles/listar_gruposF.html'

    def get_queryset(self):
        return self.model.objects.filter(iddocente__usuario =self.request.user).filter(estadoG=True).order_by('idperiodo')

    
    def get_context_data(self, **kwargs):
        contexto = {}

        list=[]
        for grupo in self.get_queryset():
            var = Det_Grupo.objects.filter(idgrupo =grupo.idgrupo).count()
            list.append({'grupos':grupo,'counter':var})
        contexto['var']=list

        return contexto
    def get(self, request, *args, **kwargs):    
        return render(request, self.template_name, self.get_context_data())

class CalificarAdmin(LoginYSuperUsuarioMixin, View):
    model = Grupo
    form_class = GruposForm
    template_name = 'ingles/calif_admin.html'

    def get_queryset(self):
        return self.model.objects.filter(estadoG=True).order_by('idperiodo')

    
    def get_context_data(self, **kwargs):
        contexto = {}

        list=[]
        for grupo in self.get_queryset():
            var = Det_Grupo.objects.filter(idgrupo =grupo.idgrupo).count()
            list.append({'grupos':grupo,'counter':var})
        contexto['var']=list

        return contexto
    def get(self, request, *args, **kwargs):    
        return render(request, self.template_name, self.get_context_data())

class CalificarGrupo(LoginRequiredMixin,LoginYFacilitador, View):
    model = Grupo
    form_class = GruposForm
    template_name = 'ingles/califi_grupos.html'

    def get_queryset(self):
        return self.model.objects.filter(iddocente__usuario =self.request.user).filter(estadoG=True).order_by('idmateria')

    
    def get_context_data(self, **kwargs):
        contexto = {}

        list=[]
        for grupo in self.get_queryset():
            var = Det_Grupo.objects.filter(idgrupo =grupo.idgrupo).count()
            list.append({'grupos':grupo,'counter':var})
        contexto['var']=list

        return contexto
    def get(self, request, *args, **kwargs):    
        return render(request, self.template_name, self.get_context_data())

class ActualizarGrupo(LoginYSuperUsuarioMixin, UpdateView):
    model = Grupo
    form_class = GruposForm
    template_name = 'ingles/grupo.html'

    def post (self, request, *arg, **kwargs):
        print('estas en post')
        form = self.form_class(request.POST, instance= self.get_object())
        print(form)
        if form.is_valid():
            print('Si te dejo entrar al form.is_valid()')
            grup = form.save(commit = False)
            grup.estado = True
            # print(grup.estado)
            grup.save()
        return redirect('curso:listar_grupos')

class CrearGrupo(LoginYSuperUsuarioMixin, CreateView):
    model =Grupo
    template_name = 'ingles/crear_grupos.HTML'
    form_class = GruposForm
    success_url = reverse_lazy('curso:listar_grupos')

class EliminarGrupo(LoginYSuperUsuarioMixin, DeleteView):
    model = Grupo
    #success_url = reverse_lazy('curso:listar_grupos')
    def post(self,request,pk, *args, **kwargs):
        object = Grupo.objects.get(idgrupo = pk)
        object.estado = False
        object.save()
        return redirect('curso:listar_grupos')
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""

""" CRUD DET_GRUPO """

class ListadoGrupoDetalle(LoginRequiredMixin, ListView):
    model = Det_Grupo
    template_name = 'ingles/listar_gruposdetalle.html'
    context_object_name = 'gruposdetalles'

class ActualizarGrupoDetalle(LoginYSuperUsuarioMixin, UpdateView):
    model = Det_Grupo
    template_name = 'ingles/crear_gruposdetalle.html'
    form_class = GruposDetalleForm
    success_url = reverse_lazy('curso:calificar_gruposF')

class CrearGrupoDetalle(LoginYSuperUsuarioMixin, CreateView):
    model = Det_Grupo
    template_name = 'ingles/crear_gruposdetalle.html'
    form_class = GruposDetalleForm
    success_url = reverse_lazy('curso:listar_grupos')

class CrearGrupoDetalleAdm(LoginYSuperUsuarioMixin, CreateView):
    model = Det_Grupo
    template_name = 'ingles/crear_gruposdetalle.html'
    form_class = GruposDetalleAdmForm
    success_url = reverse_lazy('curso:listar_grupos')

class EliminarGrupoDetalle(LoginYSuperUsuarioMixin, DeleteView):
    model = Det_Grupo
    success_url = reverse_lazy('curso:listar_grupos')

def calificar(request, id):
    title = "Calificar"
    grupodetFormset = modelformset_factory(Det_Grupo, form=GruposDetalleFormset, extra=0)
    queryset = Det_Grupo.objects.filter(idgrupo = id).filter(foliopago__idestado = 2).order_by('idestudiante__apellidop')
    gruposdetalles = Det_Grupo.objects.filter(idgrupo = id).order_by('idestudiante__apellidop')
    formset = grupodetFormset(request.POST or None,queryset=queryset )
    if formset.is_valid():
        instances = formset.save(commit=False)
        for instance in instances:
            # instance.id__idestudiante__cursando = False
            # i = Det_Grupo.objects.filter(id = instance.id)
            # instance.idestudiante.cursando
            Estudiante.objects.filter(idestudiante = instance.idestudiante.idestudiante).update(cursando = False) 
            # print(i)
            instance.save()
        if request.user.username == "Coordinador":
            return redirect('/calificar_admin')
        return redirect('/calificar_grupos')
    context = {
        "formset": formset,
        "gruposdetalles": gruposdetalles,
                }
    # print(formset.is_valid())
    return render(request,"ingles/calificar.html", context)


def listaDetalle(request, id):
    gruposdetalles = Det_Grupo.objects.filter(idgrupo = id).order_by('idestudiante__apellidop')
    var = Det_Grupo.objects.filter(idgrupo = id).count()
    varA = Det_Grupo.objects.filter(idgrupo = id).filter(foliopago__idestado = 2).count()
    context = {
                'gruposdetalles': gruposdetalles,
                'var': var,
                'varA': varA,
            }
    print(var)
    g =Grupo.objects.get(idgrupo = id).capacidad
    print(g)
    print(context)
    # if var >= Grupo.objects.get(idgrupo = id).capacidad:
    #     alumno = Grupo.objects.filter(idgrupo = id).update(estado=False)
    
    return render(request,'ingles/listar_gruposdetalle.html', {'context': context})

""" CRUD DOCENTES """
class ListadoDocente(LoginYSuperUsuarioMixin, View):
    model = Docente
    form_class = DocenteForm
    template_name = 'ingles/listar_docente.html'

    def get_queryset(self):
        return self.model.objects.filter(estado = True).order_by('-iddocente')
    
    def get_context_data(self, **kwargs):
        contexto = {}
        contexto['docentes']=self.get_queryset()
        return contexto
    
    def get(self, request, *args, **kwargs):    
        return render(request, self.template_name, self.get_context_data())
    
class ActualizarDocente(LoginYSuperUsuarioMixin, UpdateView):
    model = Docente
    form_class = FacilitadorForm
    template_name = 'ingles/docente.html'
    success_url = reverse_lazy('curso:listar_docente')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['docentes']=Docente.objects.filter(estado = True)
        return context

class CrearDocente(LoginYSuperUsuarioMixin, CreateView):
    model = Docente
    template_name = 'ingles/crear_docente.html'
    form_class = DocenteForm
    success_url = reverse_lazy('curso:listar_docente')

class EliminarDocente(LoginYSuperUsuarioMixin, DeleteView):
    model = Docente
    #success_url = reverse_lazy('curso:listar_estudiante')

    def post(self,request,pk, *args, **kwargs):
        object = Docente.objects.get(iddocente = pk)
        object.estado = False
        object.save()
        return redirect('curso:listar_docente')
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""

""" CRUD PAGO """
class ListadoPago(LoginYSuperUsuarioMixin, View):
    model = Pago
    form_class = PagoForm
    template_name = 'ingles/listar_pagos.html'
    
    def get_queryset(self):
        return self.model.objects.filter(estado = True).order_by('-foliopago')
    
    def get_context_data(self, **kwargs):
        contexto = {}
        contexto['pagos']=self.get_queryset()
        return contexto
    
    def get(self, request, *args, **kwargs):    
        return render(request, self.template_name, self.get_context_data())
    
class ListadoPagoE(LoginRequiredMixin, View):
    model = Pago
    form_class = PagoForm
    template_name = 'ingles/listar_pagos.html'
    
    def get_queryset(self):
        return self.model.objects.filter(usuario =self.request.user.username).order_by('-idmateria')
    
    def get_context_data(self, **kwargs):
        contexto = {}
        contexto['pagos']=self.get_queryset()
        return contexto
    
    def get(self, request, *args, **kwargs):    
        return render(request, self.template_name, self.get_context_data())
    
class ActualizarPago(LoginRequiredMixin, UpdateView):
    model = Pago
    form_class = PagoFormM
    template_name = 'ingles/pago.html'
    success_url = reverse_lazy('curso:listar_pagoE')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pagos']=Pago.objects.filter(estado = True)
        return context
class ActualizarPago2(LoginRequiredMixin, UpdateView):
    model = Pago
    form_class = PagoFormF
    template_name = 'ingles/pago2.html'

    def get_success_url(self):
        referer = self.request.META.get('HTTP_REFERER')
        if referer:
            return referer
        return reverse('curso:listar_financiero', args=['all'])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pagos'] = Pago.objects.filter(estado=True)
        return context

    def form_valid(self, form):
        try:
            response = super().form_valid(form)
            pago = self.object  # Obtener el objeto Pago actualizado

            if pago.idestudiante.email:
                try:
                    # Preparar el contenido del correo electrónico
                    subject = 'Estado de pago | Coordinación de Lenguas Extranjeras | Departamento de Recursos Financieros'
                    message_html = render_to_string('ingles/pago_correo.html', {'object': pago})
                    message_plain = strip_tags(message_html)

                    # Enviar el correo electrónico
                    send_mail(
                        subject,
                        message_plain,
                        'soportecleita@gmail.com',  # Remitente del correo
                        [pago.idestudiante.email],  # Lista de destinatarios
                        html_message=message_html,
                    )
                    messages.success(self.request, 'El correo electrónico se ha enviado correctamente.')
                except Exception as e:
                    # Manejar errores de envío de correo
                    messages.error(self.request, f'Error al enviar el correo electrónico: {e}')
            else:
                # Agregar un mensaje de alerta si no hay correo electrónico del estudiante
                messages.warning(self.request, 'El correo electrónico no se ha enviado porque el estudiante no tiene una dirección de correo electrónico registrada.')

            # Devolver una respuesta JSON indicando éxito
            return JsonResponse({'success': True})

        except Exception as e:
            # Manejar otros errores generales
            messages.error(self.request, f'Error al verificar el pago: {e}')
            return JsonResponse({'success': False, 'error_message': str(e)})


class ActualizarPago3(LoginRequiredMixin, UpdateView):
    model = Pago
    form_class = PagoFormF
    template_name = 'ingles/pago2.html'

    def get_success_url(self):
        # Retornar None para indicar que no se debe redirigir explícitamente
        return None

    def form_valid(self, form):
        try:
            response = super().form_valid(form)
            messages.success(self.request, 'Cambios realizados.')
            # Devolver una respuesta JSON indicando éxito
            return JsonResponse({'success': True})

        except Exception as e:
            messages.error(self.request, f'Error al verificar el pago: {e}')
            return JsonResponse({'error': str(e)}, status=400)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pagos'] = Pago.objects.filter(estado=True)
        return context

class ActualizarPagoPhone(LoginRequiredMixin, UpdateView):
    model = Pago
    form_class = PagoFormM
    template_name = 'ingles/pagoPhone.html'
    success_url = reverse_lazy('curso:listar_pagoE')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pagos']=Pago.objects.filter(estado = True)
        return context

class CrearPago(LoginRequiredMixin, CreateView):
    model = Pago
    template_name = 'ingles/crear_pago.html'
    form_class = PagoForm
    second_form_class = GruposDetalleForm
    success_url = reverse_lazy('curso:listar_pagoE')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pagos']=Pago.objects.filter(estado = True)
        context['estudiante']= Estudiante.objects.get(usuario = self.request.user)  
        if 'form' != context:
            context['form'] = self.form_class(self.request.GET)
        if 'form2' != context:
            context ['form2'] = self.second_form_class(self.request.GET)
        return context

    def post (self, request, *arg, **kwargs):
        self.object = self.get_object
        form = self.form_class(request.POST, request.FILES or None)
        form2 = self.second_form_class(request.POST)
        if form.is_valid() and form2.is_valid():
            # Obtener el estudiante actual
            estudiante = Estudiante.objects.get(usuario=request.user)
            periodo = form.cleaned_data['idperiodo']

            # Verificar si ya existe un pago para este estudiante y periodo
            if Pago.objects.filter(idestudiante=estudiante, idperiodo=periodo).exists():
                messages.warning(request, 'Ya existe un pago registrado para este periodo.')
                return HttpResponseRedirect(reverse_lazy('curso:listar_pagoE'))  # Redirigir a la página del formulario
            
            pago = form.save(commit = False)
            grup = form2.save(commit = False)
            pago.idestudiante = Estudiante.objects.get(usuario = self.request.user) 
            grup.idestudiante = pago.idestudiante
            pago.usuario = self.request.user.username
            grup.calif = ""
            grup.idperiodo = pago.idperiodo
            grup.idgrupo = pago.idgrupo
            pago.pagocurso = request.FILES['pagocurso']
            print(pago.pagocurso)
            pago.save()
            print(pago.foliopago)
            grup.foliopago = Pago.objects.get(foliopago = pago.foliopago) 
            grup.save()
            Estudiante.objects.filter(usuario = self.request.user).update(cursando = True)
            var = Det_Grupo.objects.filter(idgrupo = grup.idgrupo.idgrupo).count()
            print(var)
            g =Grupo.objects.get(idgrupo = grup.idgrupo.idgrupo).capacidad
            print(g)
            if var >= Grupo.objects.get(idgrupo = grup.idgrupo.idgrupo).capacidad:
                Grupo.objects.filter(idgrupo = grup.idgrupo.idgrupo).update(estado=False)
            return HttpResponseRedirect(reverse_lazy('curso:listar_pagoE'))
        else:
            return self.render_to_response(self.get_context_data(form = form, form2 = form2))



#############################################################################################################3
class CrearPago2(LoginRequiredMixin, CreateView):
    model = Pago
    template_name = 'ingles/crear_pago2.html'
    form_class = PagoForm
    second_form_class = GruposDetalleForm
    success_url = reverse_lazy('curso:listar_pagoE')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['pagos']=Pago.objects.filter(estado = True)
        context['estudiante']= Estudiante.objects.get(usuario = self.request.user)  
        if 'form' != context:
            context['form'] = self.form_class(self.request.GET)
        if 'form2' != context:
            context ['form2'] = self.second_form_class(self.request.GET)
        return context

    def post (self, request, *arg, **kwargs):
        self.object = self.get_object
        form = self.form_class(request.POST, request.FILES or None)
        form2 = self.second_form_class(request.POST)
        if form.is_valid() and form2.is_valid():
            # Obtener el estudiante actual
            estudiante = Estudiante.objects.get(usuario=request.user)
            periodo = form.cleaned_data['idperiodo']

            # Verificar si ya existe un pago para este estudiante y periodo
            if Pago.objects.filter(idestudiante=estudiante, idperiodo=periodo).exists():
                messages.warning(request, 'Ya existe un pago registrado para este periodo.')
                return HttpResponseRedirect(reverse_lazy('curso:listar_pagoE'))  # Redirigir a la página del formulario
            
            pago = form.save(commit = False)
            grup = form2.save(commit = False)
            pago.idestudiante = Estudiante.objects.get(usuario = self.request.user) 
            grup.idestudiante = pago.idestudiante
            pago.usuario = self.request.user.username
            grup.calif = ""
            grup.idperiodo = pago.idperiodo
            grup.idgrupo = pago.idgrupo
            pago.pagocurso = request.FILES['pagocurso']
            print(pago.pagocurso)
            pago.save()
            print(pago.foliopago)
            grup.foliopago = Pago.objects.get(foliopago = pago.foliopago) 
            grup.save()
            Estudiante.objects.filter(usuario = self.request.user).update(cursando = True)
            var = Det_Grupo.objects.filter(idgrupo = grup.idgrupo.idgrupo).count()
            print(var)
            g =Grupo.objects.get(idgrupo = grup.idgrupo.idgrupo).capacidad
            print(g)
            if var >= Grupo.objects.get(idgrupo = grup.idgrupo.idgrupo).capacidad:
                Grupo.objects.filter(idgrupo = grup.idgrupo.idgrupo).update(estado=False)
            return HttpResponseRedirect(reverse_lazy('curso:listar_pagoE'))
        else:
            return self.render_to_response(self.get_context_data(form = form, form2 = form2))

# AJAX
def load_cities(request):
    idperiodo = request.GET.get('idperiodo')
    idmateria = request.GET.get('idmateria')
    grupos = Grupo.objects.filter(idperiodo=idperiodo).filter(idmateria=idmateria).filter(estado=True)
    return render(request,'ingles/city_dropdown_list_options.html', {'grupos': grupos})

#############################################################################################################

class EliminarPago(LoginYSuperUsuarioMixin, DeleteView):
    model = Pago
    #success_url = reverse_lazy('curso:listar_estudiante')

    def post(self,request,pk, *args, **kwargs):
        object = Pago.objects.get(foliopago = pk)
        object.estado = False
        object.save()
        return redirect('curso:listar_pagoE')
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""

""" CRUD PERIODO """
class ListadoPeriodo(LoginYSuperUsuarioMixin, View):
    model = Periodo
    form_class = PeriodoForm
    template_name = 'ingles/listar_periodo.html'

    def get_queryset(self):
        return self.model.objects.filter(estado = True)
    
    def get_context_data(self, **kwargs):
        contexto = {}
        contexto['periodos']=self.get_queryset()
        return contexto
    
    def get(self, request, *args, **kwargs):    
        return render(request, self.template_name, self.get_context_data())

class ActualizarPeriodo(LoginYSuperUsuarioMixin, UpdateView):
    model = Periodo
    form_class = PeriodoForm
    template_name = 'ingles/periodo.html'
    success_url = reverse_lazy('curso:listar_periodo')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['periodos']=Periodo.objects.filter(estado = True)
        return context

class CrearPeriodo(LoginYSuperUsuarioMixin, CreateView):
    model = Periodo
    template_name = 'ingles/crear_periodo.html'
    form_class = PeriodoForm
    success_url = reverse_lazy('curso:listar_periodo')

class EliminarPeriodo(LoginYSuperUsuarioMixin, DeleteView):
    model = Periodo
    #success_url = reverse_lazy('curso:listar_estudiante')

    def post(self,request,pk, *args, **kwargs):
        object = Periodo.objects.get(idperiodo = pk)
        object.estado = False
        object.save()
        return redirect('curso:listar_periodo')
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
class Pdf(View):
    def get(self, request, *args, **kwargs):
        try:
            template = get_template('layouts/ListaE.html')
            context = {
                    'sale': Grupo.objects.get(idgrupo=self.kwargs['pk']),
                    'comp': {'name': 'curso de ingles', 'ruc': '999999', 'address':'acapulco'}
                    # 'comp': Det_Grupo.objects.filter(idgrupo = self.kwargs['pk']).count(),
            }
            html = template.render(context)
            response = HttpResponse(content_type='application/pdf')
            # response['Content-Disposition'] = 'attachment; filename="report.pdf"'
            pisaStatus = pisa.CreatePDF(
                html, dest=response)
            # pdf=render_to_pdf(template_name,context)
            print(context)
            return response
        except:
            pass
        return HttpResponseRedirect(reverse_lazy('curso:listar_grupos'))
        
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""

class sale(LoginRequiredMixin,View):
    def get(self, request,*args,**kwargs):
        try:
            template = get_template('layouts/ListaC.html')
            context = {
                'sale': Det_Grupo.objects.filter(idgrupo=self.kwargs['pk']).order_by('idestudiante__apellidop'),
                'comp': Det_Grupo.objects.filter(idgrupo = self.kwargs['pk']).count(),
                'time': datetime.date(datetime.now()),
                'ico' : '{}{}'.format(settings.MEDIA_URL, 'TecNMHorizontal1.png')
            }
            html = template.render(context)
            css_url = os.path.join(settings.BASE_DIR, 'static/css/bootstrap.min.css')
            pdf = HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(stylesheets=[CSS(css_url)])
            return HttpResponse(pdf, content_type='application/pdf')
        except:
            pass
        return HttpResponseRedirect(reverse_lazy('ingles:calificar_grupos'))

""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
class saleF(LoginRequiredMixin,View):
    def get(self, request,*args,**kwargs):
        try:
            template = get_template('layouts/ListaE.html')
            context = {
                'sale': Det_Grupo.objects.filter(idgrupo=self.kwargs['pk']).order_by('idestudiante__apellidop'),
                'comp': Det_Grupo.objects.filter(idgrupo = self.kwargs['pk']).count(),
                'ico' : '{}{}'.format(settings.MEDIA_URL, 'TecNMHorizontal1.png')
            }
            html = template.render(context)
            css_url = os.path.join(settings.BASE_DIR, 'static/css/bootstrap.min.css')
            pdf = HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(stylesheets=[CSS(css_url)])
            return HttpResponse(pdf, content_type='application/pdf')
        except:
            pass
        return HttpResponseRedirect(reverse_lazy('ingles:listar_gruposF'))

class saleP(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        try:
            template = get_template('layouts/ListaP.html') 
            context = {
                'sale': Det_Grupo.objects.filter(idgrupo=self.kwargs['pk']).order_by('idestudiante__apellidop'),
                'comp': Det_Grupo.objects.filter(idgrupo=self.kwargs['pk']).count(),
                'ico': '{}{}'.format(settings.MEDIA_URL, 'TecNMHorizontal2.png')  
            }
            html = template.render(context)
            css_url = os.path.join(settings.BASE_DIR, 'static/css/bootstrap.min.css')
            pdf = HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(stylesheets=[CSS(css_url)])
            return HttpResponse(pdf, content_type='application/pdf')
        except:
            pass
        return HttpResponseRedirect(reverse_lazy('ingles:listar_gruposP'))

""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
class Excel(LoginRequiredMixin,TemplateView):
    def get(self, request,*args,**kwargs):
        sale = Det_Grupo.objects.filter(idgrupo=self.kwargs['pk']).order_by('idestudiante__apellidop')
        comp = Det_Grupo.objects.filter(idgrupo = self.kwargs['pk']).count()
        wb = Workbook()
        ws = wb.active
        title = ""

        # Creamos un objeto de imagen con la imagen convertida
        img = Image('/var/www/html/CursoIngles/imagen_convertida.png')
        # Definimos la celda donde queremos insertar la imagen
        celda = ws['C4']

        # Añadimos la imagen a la hoja de cálculo en la celda especificada
        ws.add_image(img, celda.coordinate)
        
        # Cambiamos los márgenes de la página en la hoja activa
        ws.page_margins.top = 0.75  # Márgen superior en pulgadas
        ws.page_margins.bottom = 0.75  # Márgen inferior en pulgadas
        ws.page_margins.left = 0.25  # Márgen izquierdo en pulgadas
        ws.page_margins.right = 0.25  # Márgen derecho en pulgadas
        ws.page_margins.header = 0.29  # Márgen de encabezado en pulgadas
        ws.page_margins.footer = 0.29  # Márgen de pie de página en pulgadas


        ws.column_dimensions['A'].width = 2
        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['O'].width = 5
        ws.column_dimensions['I'].width = 4
        ws.column_dimensions['J'].width = 4
        ws.column_dimensions['L'].width = 4
        ws.column_dimensions['N'].width = 4
        ws.column_dimensions['O'].width = 4
        # Borders
        ws.merge_cells('B2:O2')
        ws['B2'].border = Border(left=Side(style='medium'), 
               top=Side(style='medium'))
        ws['C2'].border = Border(top=Side(style='medium'))
        ws['D2'].border = Border(top=Side(style='medium'))
        ws['E2'].border = Border(top=Side(style='medium'))
        ws['F2'].border = Border(top=Side(style='medium'))
        ws['G2'].border = Border(top=Side(style='medium'))
        ws['H2'].border = Border(top=Side(style='medium'))
        ws['I2'].border = Border(top=Side(style='medium'))
        ws['J2'].border = Border(top=Side(style='medium'))
        ws['K2'].border = Border(top=Side(style='medium'))
        ws['L2'].border = Border(top=Side(style='medium'))
        ws['M2'].border = Border(top=Side(style='medium'))
        ws['N2'].border = Border(top=Side(style='medium'))
        ws['O2'].border = Border(right=Side(style='medium'), 
               top=Side(style='medium'))
        ws['B3'].border = Border(left=Side(style='medium'))
        ws['B4'].border = Border(left=Side(style='medium'))
        ws['B5'].border = Border(left=Side(style='medium'))
        ws['B6'].border = Border(left=Side(style='medium'))
        ws['B7'].border = Border(left=Side(style='medium'))
        ws['B8'].border = Border(left=Side(style='medium'))
        ws['B9'].border = Border(left=Side(style='medium'))
        ws['B10'].border = Border(left=Side(style='medium'), 
               bottom=Side(style='medium'))
        ws['C10'].border = Border(bottom=Side(style='medium'))
        ws['D10'].border = Border(bottom=Side(style='medium'))
        ws['E10'].border = Border(bottom=Side(style='medium'))
        ws['F10'].border = Border(bottom=Side(style='medium'))
        ws['G10'].border = Border(bottom=Side(style='medium'))
        ws['H10'].border = Border(bottom=Side(style='medium'))
        ws['I10'].border = Border(bottom=Side(style='medium'))
        ws['J10'].border = Border(bottom=Side(style='medium'))
        ws['K10'].border = Border(bottom=Side(style='medium'))
        ws['L10'].border = Border(bottom=Side(style='medium'))
        ws['M10'].border = Border(bottom=Side(style='medium'))
        ws['N10'].border = Border(bottom=Side(style='medium'))
        ws['O10'].border = Border(right=Side(style='medium'), 
               bottom=Side(style='medium'))
        ws['O3'].border = Border(right=Side(style='medium'))
        ws['O4'].border = Border(right=Side(style='medium'))
        ws['O5'].border = Border(right=Side(style='medium'))
        ws['O6'].border = Border(right=Side(style='medium'))
        ws['O7'].border = Border(right=Side(style='medium'))
        ws['O8'].border = Border(right=Side(style='medium'))
        ws['O9'].border = Border(right=Side(style='medium'))        
        # End Borders

        ws['C3'].font = Font(bold=True, size=18)
        ws['C3'] = 'Tecnológico Nacional de México Campus Acapulco'
        ws['C4'].font = Font(size=15)
        ws['C4'] = 'Coordinación de Lenguas Extranjeras'
        ws['C5'].font = Font(bold=True)
        ws['C5'] = 'ACTA DE CALIFICACIONES'

        ws.merge_cells('C3:N3')
        ws.merge_cells('C4:N4')
        ws.merge_cells('C5:N5')
        ws['C3'].alignment = Alignment(horizontal='center')
        ws['C4'].alignment = Alignment(horizontal='center')
        ws['C5'].alignment = Alignment(horizontal='center')

        con = 1
        cont = 18
        for grupodetalle in sale:
            ws['C7'].font = Font(bold=True)
            ws['C7'] = 'NOMBRE DEL FACILITADOR:'
            ws.merge_cells('C7:E7')
            ws['F7'] = (str(grupodetalle.idgrupo.iddocente.nombre or '') + ' ' +
            str(grupodetalle.idgrupo.iddocente.apellidop or '') + ' ' +
            str(grupodetalle.idgrupo.iddocente.apellidom or ''))
            ws['C8'].font = Font(bold=True)
            ws['C8'] = 'PERIODO:'
            ws['D8'] = grupodetalle.idperiodo.periodo
            ws['C9'].font = Font(bold=True)
            ws['C9'] = 'NIVEL:'
            ws['D9'] = grupodetalle.foliopago.idmateria.nombremateria

            ws['L8'].font = Font(bold=True)
            ws['L8'] = 'GRUPO:'
            ws['L8'].alignment = Alignment(horizontal='right')
            ws['M8'] = grupodetalle.idgrupo.grupo
            ws['M8'].alignment = Alignment(horizontal='center')

            ws['L9'].font = Font(bold=True)
            ws['L9'] = 'TOTAL DE ALUMNOS:'
            ws['L9'].alignment = Alignment(horizontal='right')
            ws['M9'] = comp
            ws['M9'].alignment = Alignment(horizontal='center')

            title = grupodetalle.idperiodo.periodo+' '+grupodetalle.foliopago.idmateria.nombremateria+' GRUPO '+grupodetalle.idgrupo.grupo

            #Horario

            ws['B12'].font = Font(bold=True)
            ws['B12'] = 'HORARIO'
            ws.merge_cells('B12:O12')
            ws['B12'].alignment = Alignment(horizontal='center')
            ws['B13'] = grupodetalle.idgrupo.idmodalidad.modalidad
            ws.merge_cells('B13:O13')
            ws['B13'].alignment = Alignment(horizontal='center')

            ws['B14'] = 'HORA'
            ws.merge_cells('B14:G14')
            ws['B14'].alignment = Alignment(horizontal='center')
            ws['H14'] = 'AULA'
            ws.merge_cells('H14:O14')
            ws['H14'].alignment = Alignment(horizontal='center')

            ws['B15'] = grupodetalle.idgrupo.horario
            ws.merge_cells('B15:G15')
            ws['B15'].alignment = Alignment(horizontal='center')
            ws['H15'] = grupodetalle.idgrupo.idaula.nombre if grupodetalle.idgrupo.idaula and grupodetalle.idgrupo.idaula.nombre is not None else ''
            ws.merge_cells('H15:O15')
            ws['H15'].alignment = Alignment(horizontal='center')



            #End Horario    




            ws['B17'].font = Font(bold=True)
            ws['B17'] = 'No.'
            ws['C17'].font = Font(bold=True)
            ws['C17'] = 'No. CONTROL'
            ws.merge_cells('C17:D17')
            ws['E17'].font = Font(bold=True)
            ws['E17'] = 'NOMBRE DEL ALUMNO'
            ws.merge_cells('E17:J17')
            ws['E17'].alignment = Alignment(horizontal='center')
            ws['K17'].font = Font(bold=True)
            ws['K17'] = 'CARRERA'
            ws.merge_cells('K17:L17')
            ws['K17'].alignment = Alignment(horizontal='center')
            ws['M17'].font = Font(bold=True)
            ws['M17'] = 'CALIFICACIÓN'
            ws.merge_cells('M17:O17')
            ws['M17'].alignment = Alignment(horizontal='center')
            color_gris = PatternFill(start_color='bbbbbb', end_color='bbbbbb', fill_type='solid')

            # Aplicamos el color de fondo a la celda A1
            ws['B12'].fill = color_gris
            ws['B13'].fill = color_gris
            ws['B17'].fill = color_gris
            ws['C17'].fill = color_gris
            ws['E17'].fill = color_gris
            ws['K17'].fill = color_gris
            ws['M17'].fill = color_gris
            
            if grupodetalle.foliopago.idestado.estado == 'Aceptado':
                ws.cell(row=cont, column=2).value = con
                ws.cell(row=cont, column=2).alignment = Alignment(horizontal='center')
                ws.cell(row=cont, column=3).value = grupodetalle.idestudiante.nocontrol
                ws.merge_cells(start_row=cont, start_column=3, end_row=cont, end_column=4)
                ws.cell(row=cont, column=3).alignment = Alignment(horizontal='center')

                try:
                    apellidop = grupodetalle.idestudiante.apellidop if grupodetalle.idestudiante.apellidop else ""
                    apellidom = grupodetalle.idestudiante.apellidom if grupodetalle.idestudiante.apellidom else ""
                    nombre = grupodetalle.idestudiante.nombre if grupodetalle.idestudiante.nombre else ""

                    full_name = (apellidop + ' ' + apellidom + ' ' + nombre).upper()
                    ws.cell(row=cont, column=5).value = full_name                    
                    ws.merge_cells(start_row=cont, start_column=5, end_row=cont, end_column=10)
                except AttributeError:
                    ws.cell(row=cont, column=5).value = None
                    ws.merge_cells(start_row=cont, start_column=5, end_row=cont, end_column=10)
                try:
                    ws.cell(row=cont, column=11).value = grupodetalle.idestudiante.idcarrera.nombrecarrera
                    ws.merge_cells(start_row=cont, start_column=11, end_row=cont, end_column=12)
                    ws.cell(row=cont, column=11).alignment = Alignment(horizontal='center')
                except AttributeError:
                    ws.cell(row=cont, column=11).value = None
                    ws.merge_cells(start_row=cont, start_column=11, end_row=cont, end_column=12)
                    ws.cell(row=cont, column=11).alignment = Alignment(horizontal='center')
                
                ws.cell(row=cont, column=13).value = grupodetalle.calif
                ws.merge_cells(start_row=cont, start_column=13, end_row=cont, end_column=15)
                ws.cell(row=cont, column=13).alignment = Alignment(horizontal='center')
                cont+=1
                con+=1

        # Aplica borders a todas las celdas de tabla
        bordeE = Border(left=Side(style='thin'), 
               right=Side(style='thin'), 
               top=Side(style='thin'), 
               bottom=Side(style='thin'))
        
        bordeC = Border(left=Side(style='medium'), 
               right=Side(style='medium'), 
               top=Side(style='medium'), 
               bottom=Side(style='medium'))
        
        fonsi = Font(size=10)
        
        for row in ws.iter_rows(min_row=12, max_row=15, min_col=2, max_col=15):
            for cell in row:
                cell.border = bordeC

        for row in ws.iter_rows(min_row=17, max_row=cont-1, min_col=2, max_col=15):
            for cell in row:
                cell.border = bordeE

        for row in ws.iter_rows(min_row=18, max_row=cont-1, min_col=2, max_col=15):
            for cell in row:
                cell.font = fonsi


        ws.cell(row=cont+2,column=2).value = 'Este documento no es válido si tiene tachaduras o enmendaduras'
        ws.merge_cells(start_row=cont+2, start_column=2, end_row=cont+2, end_column=8)

        ws.cell(row=cont+3,column=2).value = 'Acapulco, Gro., '+ datetime.datetime.now().date().strftime('%Y-%m-%d')
        ws.merge_cells(start_row=cont+3, start_column=2, end_row=cont+3, end_column=6)
        ws.cell(row=cont+3, column=2).font = Font(bold=True)

        ws.cell(row=cont+4,column=11).value = 'Nombre y Firma del facilitador: '
        ws.merge_cells(start_row=cont+4, start_column=11, end_row=cont+4, end_column=14)

        nombre_archivo = "Reporte_"+title+".xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
    
class ListadoDocente(LoginRequiredMixin, View):
    model = Docente
    modelg = Grupo
    form_class = DocenteForm
    template_name = 'ingles/listar_docentes.html'

    def get_queryset(self):
        return self.model.objects.filter(estado = True).order_by('iddocente')
    
    def get_querygrupos(self):
        return self.modelg.objects.filter(estadoG = True).order_by('idperiodo')
    
    def get_context_data(self, **kwargs):
        contexto = {}
        list=[]
        
        for docente in self.get_queryset():
            gActivos = 0
            if docente.ced != 'NA':
                var = Docente.objects.filter(iddocente = docente.iddocente).count()
                gTotales = Grupo.objects.filter(iddocente = docente.iddocente).count
                
                gruposA = Grupo.objects.filter(iddocente = docente.iddocente)
                for a in gruposA:
                    gCalif = Det_Grupo.objects.filter(idgrupo=a.idgrupo, calif__isnull=False).exclude(calif='').count()
                    if gCalif != Det_Grupo.objects.filter(idgrupo = a.idgrupo).count():
                        gActivos = gActivos + 1

                list.append({'docentes':docente,'counter':var,'grupos':gTotales, 'gruposA':gActivos})
        contexto['var']=list
        return contexto
    
    def get(self, request, *args, **kwargs):    
        return render(request, self.template_name, self.get_context_data())
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
def listadoGrupoDocente(request, id): 
    gruposDocente = Grupo.objects.filter(iddocente = id).order_by('idperiodo')
    contexto = {}
    list=[]
    for grupo in gruposDocente:
        var = Det_Grupo.objects.filter(idgrupo =grupo.idgrupo).count()
        list.append({'grupos':grupo,'counter':var})
    contexto['var']=list
    return render(request,'ingles/listar_gr_docente.html', contexto)
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
class ListadoAlumno(LoginRequiredMixin, View):
    model = Estudiante
    form_class = EstudianteForm
    template_name = 'ingles/listar_alumnos.html'

    def get_queryset(self):
        return self.model.objects.filter(estado = True).order_by('apellidop')
    
    def get_context_data(self, **kwargs):
        contexto = {}
        list=[]
        for alumno in self.get_queryset():
            var = Estudiante.objects.filter(idestudiante = alumno.idestudiante).count()
            list.append({'alumnos':alumno,'counter':var})
        contexto['var']=list

        periodos = []
        ultimoPeriodo = Periodo.objects.order_by('-idperiodo')
        for vari in ultimoPeriodo:
            periodos.append(vari)
        contexto['periodos'] = periodos

        return contexto
    
    def get(self, request, *args, **kwargs):    
        return render(request, self.template_name, self.get_context_data())
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
def listarHistorial(request, id):  
    alumnos = Estudiante.objects.filter(nocontrol = id)
    contexto = {}
    list = []
    for alumno in alumnos:
        gTotal = Det_Grupo.objects.filter(idestudiante = alumno.idestudiante)
        for grupo in gTotal:
            try:
                calif = int(grupo.calif)
                if calif < 70:
                    grupo.aprobado = 2
                elif calif > 100:
                        grupo.aprobado = 4
                else:
                    grupo.aprobado = 1
            except (ValueError, TypeError):
                if grupo.calif == 'NA':
                    grupo.aprobado = 2
                else:
                    grupo.aprobado = 3         
            list.append({'grupos': grupo})

            
    contexto['estudiantes']=alumnos
    contexto['grupos'] = list

    return render(request,'ingles/listar_historial.html', contexto)
""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
class ListadoFinanciero(View):
    model = Pago
    form_class = PagoForm
    template_name = 'ingles/listar_financiero.html'

    def get(self, request, pk, *args, **kwargs):
        contexto = {}
        if pk == 'default':
            ultimo_periodo = Periodo.objects.latest('idperiodo')
            contexto['pagos'] = self.model.objects.filter(idperiodo=ultimo_periodo).order_by('-fechasist')
            contexto['titulo'] = ultimo_periodo.periodo
        else:
            valores = pk.split('&')

            if valores[2] != 'None':
                if valores[2] == '1':
                    stateName = 'Pendiente'
                elif valores[2] == '2':
                    stateName = 'Aceptado'
                else:
                    stateName = 'No aceptado'
            
            pagos_query = None
            nombre_periodo = None
            titulo = None


            # Caso 1: Ambos valores no son 'None'
            if valores[1] != 'None' and valores[2] != 'None':
                if valores[1] == 'all' and valores[2] == 'all':
                    pagos_query = self.model.objects.order_by('-fechasist')
                    titulo = "Todos los períodos"
                elif valores[1] == 'all' and valores[2] != 'None':
                    pagos_query = self.model.objects.filter(idestado=valores[2]).order_by('-fechasist')
                    titulo = f"Todos los períodos | {stateName}"
                elif valores[2] == 'all' and valores[1] != 'None':
                    pagos_query = self.model.objects.filter(idperiodo=valores[1]).order_by('-fechasist')
                    if pagos_query.exists():
                        primer_pago = pagos_query.first()
                        nombre_periodo = primer_pago.idperiodo.periodo
                        titulo = f"{nombre_periodo}"
                else:
                    pagos_query = self.model.objects.filter(idperiodo=valores[1], idestado=valores[2]).order_by('-fechasist')
                    if pagos_query.exists():
                        primer_pago = pagos_query.first()
                        nombre_periodo = primer_pago.idperiodo.periodo
                        titulo = f"{nombre_periodo} | {stateName}"

            # Caso 2: Solo el valor de 'idperiodo' no es 'None'
            elif valores[1] != 'None':
                if valores[1] != 'all':
                    pagos_query = self.model.objects.filter(idperiodo=valores[1]).order_by('-fechasist')
                else:
                    pagos_query = self.model.objects.order_by('-fechasist')
                if pagos_query.exists():
                    primer_pago = pagos_query.first()
                    nombre_periodo = primer_pago.idperiodo.periodo
                    titulo = f"{nombre_periodo}"

            # Caso 3: Solo el valor de 'idestado' no es 'None'
            elif valores[2] != 'None':
                if valores[2] != 'all':
                    pagos_query = self.model.objects.filter(idperiodo=valores[2]).order_by('-fechasist')
                else:
                    pagos_query = self.model.objects.order_by('-fechasist')
                titulo = f"Todos los períodos | {stateName}"

            # Si pagos_query existe, agregarlo al contexto
            if pagos_query and pagos_query.exists():
                contexto['pagos'] = pagos_query
                if not titulo:  # Caso cuando 'all' y 'all' no tiene resultados
                    titulo = "No hay resultados"
                contexto['titulo'] = titulo
            else: 
                contexto['pagos'] = None
                contexto['titulo'] = "No hay resultados"

        periodos = []
        ultimoPeriodo = Periodo.objects.order_by('-idperiodo')
        for vari in ultimoPeriodo:
            periodos.append(vari)
        contexto['periodos'] = periodos

        return render(request, self.template_name, contexto)
 

""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
class EliminarDocentes(LoginYSuperUsuarioMixin, DeleteView):
    model = Docente
    #success_url = reverse_lazy('curso:listar_estudiante')

    def get(self,request,pk, *args, **kwargs):
        object = Docente.objects.get(iddocente = pk)
        object.estado = False
        object.save()
        return redirect('curso:listar_docentes')

class Error404View(TemplateView):
    template_name = 'home/page-404.html' 

""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""""
#logs
from django.contrib.admin.models import LogEntry

def activity_logs(request, id_usuario=None):
    """
    Muestra los registros de actividad, filtrando opcionalmente por usuario.
    Si no se proporciona un usuario, se muestran todos los registros.
    """
    # Filtrar los registros para evitar problemas con datos corruptos
    logs = LogEntry.objects.select_related('user', 'content_type').only(
        'id', 'user_id', 'action_time', 'action_flag', 'content_type'
    ).filter(action_time__isnull=False, user_id__isnull=False).order_by('-action_time')

    if id_usuario:
        logs = logs.filter(user_id=id_usuario)

    # Contexto para el template
    contexto = {
        'logs': logs,
        'id_usuario': id_usuario,
    }
    return render(request, 'ingles/logs.html', contexto)

